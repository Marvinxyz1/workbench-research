# -*- coding: utf-8 -*-
"""
KPMG Workbench API 一覧表 Excel 生成スクリプト
"""

import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# KPMG Brand Colors
KPMG_BLUE = "005EB8"
KPMG_DARK_BLUE = "00338D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "28A745"
ORANGE = "FD7E14"

# API Data
apis = [
    {
        "category": "Azure OpenAI - Inference API",
        "name": "Chat Completion",
        "endpoint": "/genai/azure/inference/chat/completions",
        "method": "POST",
        "description": "チャット補完API。GPTモデルを使用してテキスト生成を行う",
        "test_status": "テスト済み",
        "api_version": "2024-04-01-preview",
        "test_result": "200 OK",
        "response_sample": "Response: 'こんにちは (Konnichiwa)'\nTokens: in=11, out=8"
    },
    {
        "category": "Azure OpenAI - Inference API",
        "name": "Embeddings",
        "endpoint": "/genai/azure/inference/embeddings",
        "method": "POST",
        "description": "テキスト埋め込みAPI。テキストをベクトル表現に変換する",
        "test_status": "テスト済み",
        "api_version": "2024-04-01-preview",
        "test_result": "200 OK",
        "response_sample": "Embedding dimension: 3072\n(text-embedding-3-large モデル使用)"
    },
    {
        "category": "Azure AI Translator - Text API",
        "name": "Text Translation",
        "endpoint": "/translator/azure/text/translate",
        "method": "POST",
        "description": "テキスト翻訳API。多言語間のテキスト翻訳を提供",
        "test_status": "テスト済み",
        "api_version": "3.0",
        "test_result": "200 OK",
        "response_sample": "Input: 'Hello, KPMG Workbench!'\nOutput: 'こんにちは、KPMGワークベンチの皆さん!'"
    },
    {
        "category": "Azure AI Translator - Text API",
        "name": "Language Detection",
        "endpoint": "/translator/azure/text/detect",
        "method": "POST",
        "description": "言語検出API。入力テキストの言語を自動識別",
        "test_status": "テスト済み",
        "api_version": "3.0",
        "test_result": "200 OK",
        "response_sample": "Input: 'こんにちは、KPMGワークベンチ！'\nDetected: language='ja', score=1.0"
    },
    {
        "category": "Azure AI Translator - Text API",
        "name": "Break Sentence",
        "endpoint": "/translator/azure/text/breaksentence",
        "method": "POST",
        "description": "文分割API。テキストを文単位に分割する",
        "test_status": "テスト済み",
        "api_version": "3.0",
        "test_result": "200 OK",
        "response_sample": "Input: 'How are you? I am fine. What did you do today?'\nsentLen: [13, 11, 22], language: 'en'"
    },
    {
        "category": "Azure AI Translator - Document API",
        "name": "Document Translation",
        "endpoint": "/translator/azure/document/translator/document/batches",
        "method": "POST",
        "description": "文書翻訳API。ドキュメント全体の翻訳を実行",
        "test_status": "未テスト",
        "api_version": "2024-05-01",
        "test_result": "-",
        "response_sample": "ファイルアップロードが必要"
    },
    {
        "category": "Azure AI Translator - Document API",
        "name": "Cancel Translation",
        "endpoint": "/translator/azure/document/translator/document/batches/{id}",
        "method": "DELETE",
        "description": "翻訳ジョブのキャンセル。処理中の翻訳を中止する",
        "test_status": "未テスト",
        "api_version": "2024-05-01",
        "test_result": "-",
        "response_sample": "翻訳ジョブIDが必要"
    },
    {
        "category": "Azure OpenAI - Assistants API",
        "name": "Create Assistant",
        "endpoint": "/genai/azure/openai/assistants",
        "method": "POST",
        "description": "AIアシスタント作成API。カスタムアシスタントを構築",
        "test_status": "未テスト",
        "api_version": "2024-04-01-preview",
        "test_result": "-",
        "response_sample": "アシスタント設定が必要"
    },
    {
        "category": "Azure OpenAI - Assistants API",
        "name": "Create Thread",
        "endpoint": "/genai/azure/openai/threads",
        "method": "POST",
        "description": "スレッド作成API。会話スレッドを管理する",
        "test_status": "未テスト",
        "api_version": "2024-04-01-preview",
        "test_result": "-",
        "response_sample": "アシスタントIDが必要"
    },
    {
        "category": "Azure OpenAI - Completions API",
        "name": "Chat Completions",
        "endpoint": "/genai/azure/openai/deployments/{deployment-id}/chat/completions",
        "method": "POST",
        "description": "Azure OpenAI チャット補完。デプロイメント指定版",
        "test_status": "未テスト",
        "api_version": "2024-04-01-preview",
        "test_result": "-",
        "response_sample": "deployment-id の指定が必要"
    },
    {
        "category": "Azure OpenAI - Deployments",
        "name": "List Deployments",
        "endpoint": "/genai/azure/openai/deployments",
        "method": "GET",
        "description": "デプロイメント一覧取得。利用可能なモデル一覧を表示",
        "test_status": "未テスト",
        "api_version": "2024-04-01-preview",
        "test_result": "-",
        "response_sample": "管理権限が必要な可能性"
    },
    {
        "category": "RAG - Text Chunking APIs",
        "name": "Simple Text Chunker",
        "endpoint": "/text-chunking/microsoft/simple/simple",
        "method": "POST",
        "description": "テキスト分割API。文書をチャンクに分割してRAGに活用",
        "test_status": "テスト済み",
        "api_version": "-",
        "test_result": "200 OK",
        "response_sample": "Number of chunks: 2\nFirst chunk: 'This is the first paragraph...'"
    },
    {
        "category": "RAG - Azure Extraction",
        "name": "Document Extraction",
        "endpoint": "/extraction/azure/document-intelligence",
        "method": "POST",
        "description": "文書抽出API。Azure Document Intelligenceを使用",
        "test_status": "未テスト",
        "api_version": "-",
        "test_result": "-",
        "response_sample": "ドキュメントファイルが必要"
    },
    {
        "category": "RAG - Enrichment",
        "name": "Key Phrase Extraction",
        "endpoint": "/enrichment/azure/language/key-phrases",
        "method": "POST",
        "description": "キーフレーズ抽出API。文書から重要なフレーズを抽出",
        "test_status": "未テスト",
        "api_version": "-",
        "test_result": "-",
        "response_sample": "テキストデータが必要"
    },
    {
        "category": "RAG - Enrichment",
        "name": "Document Summarization",
        "endpoint": "/enrichment/azure/language/summarization",
        "method": "POST",
        "description": "文書要約API。長い文書を自動要約する",
        "test_status": "未テスト",
        "api_version": "-",
        "test_result": "-",
        "response_sample": "長文テキストが必要"
    },
    {
        "category": "RAG - Ingestion and Retrieval",
        "name": "Document Ingestion",
        "endpoint": "/rag/ingestion",
        "method": "POST",
        "description": "文書取り込みAPI。RAGシステムに文書を登録",
        "test_status": "未テスト",
        "api_version": "-",
        "test_result": "-",
        "response_sample": "インデックス設定が必要"
    },
    {
        "category": "RAG - Ingestion and Retrieval",
        "name": "Document Retrieval",
        "endpoint": "/rag/retrieval",
        "method": "POST",
        "description": "文書検索API。RAGシステムから関連文書を取得",
        "test_status": "未テスト",
        "api_version": "-",
        "test_result": "-",
        "response_sample": "インデックスとクエリが必要"
    },
]

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "KPMG Workbench API一覧"

    # Define styles
    header_fill = PatternFill(start_color=KPMG_BLUE, end_color=KPMG_BLUE, fill_type="solid")
    header_font = Font(name="Meiryo UI", size=11, bold=True, color=WHITE)

    category_fill = PatternFill(start_color=KPMG_DARK_BLUE, end_color=KPMG_DARK_BLUE, fill_type="solid")
    category_font = Font(name="Meiryo UI", size=10, bold=True, color=WHITE)

    data_font = Font(name="Meiryo UI", size=10)
    data_font_bold = Font(name="Meiryo UI", size=10, bold=True)

    tested_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    untested_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    status_font = Font(name="Meiryo UI", size=10, bold=True, color=WHITE)

    alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Headers
    headers = ["No.", "カテゴリ", "API名", "HTTPメソッド", "エンドポイント", "APIバージョン", "説明", "テスト状況", "テスト結果", "レスポンス内容"]
    col_widths = [5, 30, 22, 12, 55, 16, 40, 12, 12, 50]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    base_url = "https://api.workbench.kpmg"

    for row_idx, api in enumerate(apis, 2):
        row_data = [
            row_idx - 1,
            api["category"],
            api["name"],
            api["method"],
            f"{base_url}{api['endpoint']}",
            api["api_version"],
            api["description"],
            api["test_status"],
            api["test_result"],
            api["response_sample"]
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.border = thin_border

            if col == 1:  # No.
                cell.alignment = center_align
            elif col == 4:  # Method
                cell.alignment = center_align
                cell.font = data_font_bold
            elif col == 6:  # API Version
                cell.alignment = center_align
            elif col == 8:  # Test Status
                cell.alignment = center_align
                cell.font = status_font
                if value == "テスト済み":
                    cell.fill = tested_fill
                else:
                    cell.fill = untested_fill
            elif col == 9:  # Test Result
                cell.alignment = center_align
                if value == "200 OK":
                    cell.font = Font(name="Meiryo UI", size=10, bold=True, color="28A745")
                else:
                    cell.font = Font(name="Meiryo UI", size=10, color="999999")
            elif col == 10:  # Response Sample
                cell.alignment = left_align
                cell.font = Font(name="Consolas", size=9)
            else:
                cell.alignment = left_align

            # Alternate row colors (except status column)
            if row_idx % 2 == 0 and col not in [8]:
                if not cell.fill or cell.fill.fill_type is None:
                    cell.fill = alt_fill

    # Freeze panes
    ws.freeze_panes = "A2"

    # Add summary section
    summary_row = len(apis) + 4

    ws.cell(row=summary_row, column=1, value="サマリー").font = Font(name="Meiryo UI", size=12, bold=True)

    tested_count = sum(1 for api in apis if api["test_status"] == "テスト済み")
    total_count = len(apis)

    summary_data = [
        (summary_row + 1, "総API数:", total_count),
        (summary_row + 2, "テスト済み:", tested_count),
        (summary_row + 3, "未テスト:", total_count - tested_count),
        (summary_row + 4, "テスト完了率:", f"{tested_count/total_count*100:.1f}%"),
    ]

    for row, label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = data_font_bold
        ws.cell(row=row, column=2, value=value).font = data_font

    # Add notes
    notes_row = summary_row + 7
    ws.cell(row=notes_row, column=1, value="備考").font = Font(name="Meiryo UI", size=12, bold=True)

    notes = [
        "• Base URL: https://api.workbench.kpmg (Developer Portal外部アクセス用)",
        "• 必須ヘッダー: Ocp-Apim-Subscription-Key, x-kpmg-charge-code",
        "• チャット/エンベディングには azureml-model-deployment ヘッダーも必要",
        "• Developer Portal URL: https://developer.australiaeast.workbench.kpmg/apis",
    ]

    for i, note in enumerate(notes):
        ws.cell(row=notes_row + 1 + i, column=1, value=note).font = data_font

    # Save
    output_path = "generated_docs/KPMG_Workbench_API一覧.xlsx"
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    return output_path

if __name__ == "__main__":
    create_excel()
