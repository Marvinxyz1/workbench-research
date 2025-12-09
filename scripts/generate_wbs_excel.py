# -*- coding: utf-8 -*-
"""
KPMG AI COE WBS Excel Generator
Creates Work Breakdown Structure with 4 phases, ~56 tasks
- Phase 列合并单元格，带背景色
- Sub-Task 列合并单元格
- 月份时间线标记
- 版本号自动递增
"""

import sys
import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Windows console UTF-8 support
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# ==================== KPMG Brand Colors ====================
PHASE_COLORS = {
    "Phase 1": "00338D",  # Dark Blue - オンボーディング
    "Phase 2": "005EB8",  # KPMG Blue - 誰でも1週間で
    "Phase 3": "7030A0",  # Purple - 海外MFナレッジ
    "Phase 4": "E91E8C",  # Magenta - スピーディな
}

OWNER_COLORS = {
    "KC": "4472C4",       # 蓝色 - KC组
    "ATH": "70AD47",      # 绿色 - ATH组
    "Joint": "7030A0",    # 紫色 - 联合任务
}

HEADER_COLOR = "00338D"

# Phase display names (for merged cells)
PHASE_DISPLAY = {
    "Phase 1": "オンボーディング\nフェーズ\n(10-11月)",
    "Phase 2": '"誰でも1週間で\n迷いなく開始"\n(12-1月)',
    "Phase 3": '"海外MF\nナレッジ探索"\n(2-3月)',
    "Phase 4": '"スピーディな\nプロダクトプロセス"\n(4-6月)',
}

# ==================== WBS Task Data ====================
WBS_TASKS = [
    # Phase 1: オンボーディングフェーズ (10-11月)
    {"wbs": "1.1", "phase": "Phase 1", "subtask": "学習・認証", "action": "Prerequisites 認証完了", "owner": "KC", "months": [10], "deliverable": "認証バッジ"},
    {"wbs": "1.1.1", "phase": "Phase 1", "subtask": "学習・認証", "action": "Developer Learning Path 完了", "owner": "KC", "months": [10, 11], "deliverable": "学習修了証"},
    {"wbs": "1.1.2", "phase": "Phase 1", "subtask": "学習・認証", "action": "Knowledge Badge 取得", "owner": "KC", "months": [11], "deliverable": "バッジ証明"},
    {"wbs": "1.1.3", "phase": "Phase 1", "subtask": "学習・認証", "action": "Tech Talks 重要回視聴", "owner": "KC", "months": [10, 11], "deliverable": "視聴記録"},
    {"wbs": "1.2", "phase": "Phase 1", "subtask": "技術評価", "action": "Workbench 環境セットアップ", "owner": "KC", "months": [10], "deliverable": "環境構築完了"},
    {"wbs": "1.2.1", "phase": "Phase 1", "subtask": "技術評価", "action": "API 機能調査（Document Translation等）", "owner": "KC", "months": [10, 11], "deliverable": "API調査レポート"},
    {"wbs": "1.2.2", "phase": "Phase 1", "subtask": "技術評価", "action": "aIQ Chat 機能検証", "owner": "KC", "months": [11], "deliverable": "検証結果"},
    {"wbs": "1.2.3", "phase": "Phase 1", "subtask": "技術評価", "action": "Agent 開発フレームワーク評価", "owner": "KC", "months": [11], "deliverable": "評価レポート"},
    {"wbs": "1.3", "phase": "Phase 1", "subtask": "課題対応", "action": "手続上の障壁把握", "owner": "KC", "months": [10], "deliverable": "課題リスト"},
    {"wbs": "1.3.1", "phase": "Phase 1", "subtask": "課題対応", "action": "APIアクセス課題解決（12/1完了）", "owner": "KC", "months": [11], "deliverable": "解決策ドキュメント"},
    {"wbs": "1.3.2", "phase": "Phase 1", "subtask": "課題対応", "action": "Global WB Community ローンチ（11/13）", "owner": "Joint", "months": [11], "deliverable": "Community稼働"},
    {"wbs": "1.4", "phase": "Phase 1", "subtask": "報告・計画", "action": "技術評価レポート作成", "owner": "KC", "months": [11], "deliverable": "評価レポート"},
    {"wbs": "1.4.1", "phase": "Phase 1", "subtask": "報告・計画", "action": "Phase 2 計画策定", "owner": "KC", "months": [11], "deliverable": "計画書"},

    # Phase 2: "誰でも1週間で迷いなく開始" (12-1月)
    {"wbs": "2.1", "phase": "Phase 2", "subtask": "KC組タスク", "action": "剩余API测试完了（Document Translation等）", "owner": "KC", "months": [12], "deliverable": "テストレポート"},
    {"wbs": "2.1.1", "phase": "Phase 2", "subtask": "KC組タスク", "action": "API测试報告総括", "owner": "KC", "months": [12], "deliverable": "総括レポート"},
    {"wbs": "2.1.2", "phase": "Phase 2", "subtask": "KC組タスク", "action": "Cookbook基礎テンプレート整備", "owner": "KC", "months": [12], "deliverable": "テンプレート"},
    {"wbs": "2.1.3", "phase": "Phase 2", "subtask": "KC組タスク", "action": "ATH組への技術サポート資料提供", "owner": "KC", "months": [12], "deliverable": "技術資料"},
    {"wbs": "2.1.4", "phase": "Phase 2", "subtask": "KC組タスク", "action": "ATH組工作成果Review", "owner": "KC", "months": [1], "deliverable": "Reviewレポート"},
    {"wbs": "2.1.5", "phase": "Phase 2", "subtask": "KC組タスク", "action": "Cookbook高級用例補充（2シナリオ）", "owner": "KC", "months": [1], "deliverable": "高級用例"},
    {"wbs": "2.1.6", "phase": "Phase 2", "subtask": "KC組タスク", "action": "Agent最佳実践ドキュメント作成", "owner": "KC", "months": [1], "deliverable": "ベストプラクティス"},
    {"wbs": "2.1.7", "phase": "Phase 2", "subtask": "KC組タスク", "action": "Phase 2 総括レポート", "owner": "KC", "months": [1], "deliverable": "総括レポート"},
    {"wbs": "2.2", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "KJ開発者Onboardingサイト設計", "owner": "ATH", "months": [12], "deliverable": "設計書"},
    {"wbs": "2.2.1", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "Onboardingサイトフロントエンド開発", "owner": "ATH", "months": [12], "deliverable": "フロントエンド"},
    {"wbs": "2.2.2", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "Onboardingサイトコンテンツ作成", "owner": "ATH", "months": [1], "deliverable": "コンテンツ"},
    {"wbs": "2.2.3", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "Agent検証環境構築", "owner": "ATH", "months": [12], "deliverable": "環境構築"},
    {"wbs": "2.2.4", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "API/Agent基礎検証テスト", "owner": "ATH", "months": [12], "deliverable": "テスト結果"},
    {"wbs": "2.2.5", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "Cookbookコア用例作成（5シナリオ）", "owner": "ATH", "months": [1], "deliverable": "コア用例"},
    {"wbs": "2.2.6", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "開発者コミュニティプラットフォーム構築", "owner": "ATH", "months": [1], "deliverable": "プラットフォーム"},
    {"wbs": "2.2.7", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "開発者コミュニティ初期コンテンツ", "owner": "ATH", "months": [1], "deliverable": "初期コンテンツ"},
    {"wbs": "2.2.8", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "Onboardingサイトローンチ＆テスト", "owner": "ATH", "months": [1], "deliverable": "サイト稼働"},
    {"wbs": "2.2.9", "phase": "Phase 2", "subtask": "ATH組タスク", "action": "開発者コミュニティ正式ローンチ", "owner": "ATH", "months": [1], "deliverable": "コミュニティ稼働"},

    # Phase 3: "海外MFナレッジ探索" (2-3月)
    {"wbs": "3.1", "phase": "Phase 3", "subtask": "POC討論", "action": "POC項目テーマブレインストーミング", "owner": "Joint", "months": [1], "deliverable": "テーマリスト"},
    {"wbs": "3.1.1", "phase": "Phase 3", "subtask": "POC討論", "action": "各方ニーズ・アイデア収集", "owner": "ATH", "months": [1], "deliverable": "ニーズリスト"},
    {"wbs": "3.1.2", "phase": "Phase 3", "subtask": "POC討論", "action": "POC項目可行性初期評価", "owner": "Joint", "months": [1], "deliverable": "評価結果"},
    {"wbs": "3.1.3", "phase": "Phase 3", "subtask": "POC討論", "action": "POC項目リスト確定（2-3件）", "owner": "Joint", "months": [1], "deliverable": "確定リスト"},
    {"wbs": "3.2", "phase": "Phase 3", "subtask": "海外MF情報収集", "action": "US/UK/AU等MFのAgent開発案例収集", "owner": "ATH", "months": [2], "deliverable": "案例リスト"},
    {"wbs": "3.2.1", "phase": "Phase 3", "subtask": "海外MF情報収集", "action": "海外MF技術ドキュメント翻訳整理", "owner": "ATH", "months": [2], "deliverable": "翻訳ドキュメント"},
    {"wbs": "3.2.2", "phase": "Phase 3", "subtask": "海外MF情報収集", "action": "知識復用フレームワーク構築", "owner": "ATH", "months": [2], "deliverable": "フレームワーク"},
    {"wbs": "3.2.3", "phase": "Phase 3", "subtask": "海外MF情報収集", "action": "海外MF技術交流会議（2回）", "owner": "KC", "months": [2], "deliverable": "会議記録"},
    {"wbs": "3.2.4", "phase": "Phase 3", "subtask": "海外MF情報収集", "action": "海外方案可行性評価レポート", "owner": "KC", "months": [2], "deliverable": "評価レポート"},
    {"wbs": "3.3", "phase": "Phase 3", "subtask": "POC開始", "action": "POC項目A開始", "owner": "ATH", "months": [3], "deliverable": "POC A進捗"},
    {"wbs": "3.3.1", "phase": "Phase 3", "subtask": "POC開始", "action": "POC項目B開始", "owner": "ATH", "months": [3], "deliverable": "POC B進捗"},
    {"wbs": "3.3.2", "phase": "Phase 3", "subtask": "POC開始", "action": "POC技術サポート", "owner": "KC", "months": [3], "deliverable": "技術サポート"},
    {"wbs": "3.3.3", "phase": "Phase 3", "subtask": "POC開始", "action": "POC中期レビュー", "owner": "KC", "months": [3], "deliverable": "中期レビュー"},

    # Phase 4: "スピーディなプロダクトプロセス" (4-6月)
    {"wbs": "4.1", "phase": "Phase 4", "subtask": "POC完成", "action": "POC項目A完成＆評価", "owner": "ATH", "months": [3, 4], "deliverable": "POC A完成"},
    {"wbs": "4.1.1", "phase": "Phase 4", "subtask": "POC完成", "action": "POC項目B完成＆評価", "owner": "ATH", "months": [3, 4], "deliverable": "POC B完成"},
    {"wbs": "4.2", "phase": "Phase 4", "subtask": "本番アプリ開発", "action": "本番アプリアーキテクチャ設計Review", "owner": "KC", "months": [4], "deliverable": "設計Review"},
    {"wbs": "4.2.1", "phase": "Phase 4", "subtask": "本番アプリ開発", "action": "本番アプリ開発（項目A）", "owner": "ATH", "months": [4, 5], "deliverable": "アプリA"},
    {"wbs": "4.2.2", "phase": "Phase 4", "subtask": "本番アプリ開発", "action": "本番アプリ開発（項目B）", "owner": "ATH", "months": [4, 5], "deliverable": "アプリB"},
    {"wbs": "4.2.3", "phase": "Phase 4", "subtask": "本番アプリ開発", "action": "セキュリティ・コンプライアンス審査", "owner": "KC", "months": [4], "deliverable": "審査結果"},
    {"wbs": "4.3", "phase": "Phase 4", "subtask": "リリース", "action": "UATテスト", "owner": "ATH", "months": [5], "deliverable": "テスト結果"},
    {"wbs": "4.3.1", "phase": "Phase 4", "subtask": "リリース", "action": "本番リリース準備", "owner": "ATH", "months": [5], "deliverable": "リリース準備"},
    {"wbs": "4.3.2", "phase": "Phase 4", "subtask": "リリース", "action": "本番リリース＆監視", "owner": "ATH", "months": [5, 6], "deliverable": "リリース完了"},
    {"wbs": "4.3.3", "phase": "Phase 4", "subtask": "リリース", "action": "最終発表審批", "owner": "KC", "months": [5], "deliverable": "審批完了"},
    {"wbs": "4.4", "phase": "Phase 4", "subtask": "プロセス確立", "action": "新プロセスドキュメント化", "owner": "ATH", "months": [6], "deliverable": "プロセスドキュメント"},
    {"wbs": "4.4.1", "phase": "Phase 4", "subtask": "プロセス確立", "action": "プロジェクト総括レポート", "owner": "KC", "months": [6], "deliverable": "総括レポート"},
]


# ==================== Style Functions ====================
def apply_header_style(cell):
    """Apply KPMG branded header style"""
    cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_phase_style(cell, phase):
    """Apply phase-specific style with background color"""
    color = PHASE_COLORS.get(phase, "00338D")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )


def apply_owner_style(cell, owner):
    """Apply owner/team-specific style"""
    color = OWNER_COLORS.get(owner, "808080")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_cell_style(cell):
    """Apply standard cell style"""
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_month_marker(cell, phase):
    """Apply month marker with phase color"""
    color = PHASE_COLORS.get(phase, "00338D")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )


def get_next_version(output_dir, base_name):
    """Get the next version number for the file"""
    pattern = os.path.join(output_dir, f"{base_name}_v*.xlsx")
    existing_files = glob.glob(pattern)

    if not existing_files:
        return 1

    versions = []
    for f in existing_files:
        match = re.search(r'_v(\d+)\.xlsx$', f)
        if match:
            versions.append(int(match.group(1)))

    return max(versions) + 1 if versions else 1


def month_to_col(m):
    """Convert month number to column index (10月=F column=6)"""
    if m >= 10:
        return 6 + (m - 10)  # 10月=6(F), 11月=7(G), 12月=8(H)
    else:
        return 9 + (m - 1)   # 1月=9(I), 2月=10(J), ..., 6月=14(N)


# ==================== Main Function ====================
def main():
    """Main function to generate the WBS Excel file"""
    print("=" * 60)
    print("KPMG AI COE WBS Excel Generator")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "WBS Overview"

    # Headers
    headers = ["Phase", "WBS", "Sub-Task", "Action", "Owner",
               "10月", "11月", "12月", "1月", "2月", "3月", "4月", "5月", "6月",
               "Deliverable", "Status"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # Column widths
    widths = {'A': 20, 'B': 10, 'C': 18, 'D': 42, 'E': 10,
              'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 5, 'L': 5, 'M': 5, 'N': 5,
              'O': 25, 'P': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Track groups for merging
    phase_groups = {}
    subtask_groups = {}

    # Write data
    for row_idx, task in enumerate(WBS_TASKS, 2):
        phase = task["phase"]
        subtask = task["subtask"]

        # Phase column
        phase_cell = ws.cell(row=row_idx, column=1, value=PHASE_DISPLAY.get(phase, phase))
        apply_phase_style(phase_cell, phase)

        if phase not in phase_groups:
            phase_groups[phase] = {"start": row_idx, "end": row_idx}
        else:
            phase_groups[phase]["end"] = row_idx

        # Other columns
        ws.cell(row=row_idx, column=2, value=task["wbs"])
        apply_cell_style(ws.cell(row=row_idx, column=2))

        ws.cell(row=row_idx, column=3, value=subtask)
        apply_cell_style(ws.cell(row=row_idx, column=3))

        key = f"{phase}|{subtask}"
        if key not in subtask_groups:
            subtask_groups[key] = {"start": row_idx, "end": row_idx}
        else:
            subtask_groups[key]["end"] = row_idx

        ws.cell(row=row_idx, column=4, value=task["action"])
        apply_cell_style(ws.cell(row=row_idx, column=4))

        owner_cell = ws.cell(row=row_idx, column=5, value=task["owner"])
        apply_owner_style(owner_cell, task["owner"])

        # Month markers
        for m in task["months"]:
            col = month_to_col(m)
            cell = ws.cell(row=row_idx, column=col)
            apply_month_marker(cell, phase)

        # Empty month cells with border
        for m in list(range(10, 13)) + list(range(1, 7)):
            col = month_to_col(m)
            if m not in task["months"]:
                ws.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )

        # Deliverable and Status
        ws.cell(row=row_idx, column=15, value=task["deliverable"])
        apply_cell_style(ws.cell(row=row_idx, column=15))

        ws.cell(row=row_idx, column=16, value="未開始")
        apply_cell_style(ws.cell(row=row_idx, column=16))

    # Merge Phase cells
    for phase, indices in phase_groups.items():
        if indices["start"] != indices["end"]:
            ws.merge_cells(start_row=indices["start"], start_column=1,
                          end_row=indices["end"], end_column=1)

    # Merge Sub-Task cells
    for key, indices in subtask_groups.items():
        if indices["start"] != indices["end"]:
            ws.merge_cells(start_row=indices["start"], start_column=3,
                          end_row=indices["end"], end_column=3)

    # Freeze panes
    ws.freeze_panes = 'F2'

    # Row height
    ws.row_dimensions[1].height = 30
    for row in range(2, len(WBS_TASKS) + 2):
        ws.row_dimensions[row].height = 22

    # Save
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generated_docs")
    os.makedirs(output_dir, exist_ok=True)

    # Version management
    base_name = "KPMG_AI_COE_WBS"
    version = get_next_version(output_dir, base_name)

    output_path = os.path.join(output_dir, f"{base_name}_v{version}.xlsx")
    wb.save(output_path)

    print(f"\n✅ WBS Excel 已生成: {output_path}")
    print(f"   - 版本: v{version}")
    print(f"   - 任务数: {len(WBS_TASKS)} 项")
    print(f"   - Phase 数: {len(phase_groups)} 个")
    print("=" * 60)

    return output_path


if __name__ == "__main__":
    main()
