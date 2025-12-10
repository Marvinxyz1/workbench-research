# -*- coding: utf-8 -*-
"""
KPMG Workbench 開発スケジュール Excel 生成ツール v2
ガントチャート、タスク詳細、休暇カレンダー、マイルストーンを含む完全なスケジュール表を生成
- Phase列のセル結合と背景色
- バージョン番号自動インクリメント（既存ファイルを上書きしない）
"""

import sys
import os
import re
import glob
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# Windows console UTF-8 support
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# ==================== KPMG ブランドカラー ====================
KPMG_BLUE = "005EB8"
KPMG_DARK_BLUE = "00338D"
KC_COLOR = "4472C4"      # 青色 - KC組
ALH_COLOR = "70AD47"     # 緑色 - ALH組
JOINT_COLOR = "7030A0"   # 紫色 - 連合タスク (Joint)
HOLIDAY_COLOR = "D9D9D9" # グレー - 休暇
FREEZE_COLOR = "D9D9D9"  # グレー - Freeze Period
MAGENTA_COLOR = "E91E8C" # マゼンタ - Phase 4
DECISION_COLOR = "FFCCCC"  # 淡赤 - Go/No-Go 決定点

# Phase カラー
PHASE_COLORS = {
    "Phase 1": KPMG_DARK_BLUE,  # 濃紺 - オンボーディング
    "Phase 2": KPMG_BLUE,       # 青 - 誰でも1週間で
    "Phase 3": ALH_COLOR,       # 緑 - 海外MFナレッジ
    "Phase 4": MAGENTA_COLOR,   # マゼンタ - スピーディな
}

# Phase display names (for merged cells)
PHASE_DISPLAY = {
    "Phase 1": "オンボーディング\nフェーズ\n(10-11月)",
    "Phase 2": "誰でも1週間で\n迷いなく開始\n(12-1月)",
    "Phase 3": "海外MF\nナレッジ探索\n(2-3月)",
    "Phase 4": "スピーディな\nプロダクトプロセス\n(4-6月)",
}

# ステータスカラー（タスクデータのstatus値と一致）
STATUS_COLORS = {
    "未開始": "BFBFBF",    # グレー
    "進行中": "FFC000",    # 黄色
    "完了": "92D050",      # 緑色
    "ブロック": "FF0000",  # 赤色
}

# ==================== タスクデータ (WBS) ====================
# 新構造: category（機能別分類）, start_date/end_date（精確日期）, is_freeze（凍結期間フラグ）
ALL_TASKS = [
    # Phase 1: オンボーディングフェーズ (10-11月) - 既存データ維持（新フィールド追加）
    {"wbs": "1.1", "phase": "Phase 1", "category": "学習・認証", "name": "Prerequisites 認証完了", "team": "Joint", "start_date": "2025-10-01", "end_date": "2025-10-31", "deliverable": "認証バッジ", "status": "完了", "is_freeze": False},
    {"wbs": "1.1.1", "phase": "Phase 1", "category": "学習・認証", "name": "Developer Learning Path 完了", "team": "Joint", "start_date": "2025-10-01", "end_date": "2025-11-30", "deliverable": "学習修了証", "status": "完了", "is_freeze": False},
    {"wbs": "1.1.2", "phase": "Phase 1", "category": "学習・認証", "name": "Knowledge Badge 取得", "team": "Joint", "start_date": "2025-11-01", "end_date": "2025-11-30", "deliverable": "バッジ証明", "status": "完了", "is_freeze": False},
    {"wbs": "1.2", "phase": "Phase 1", "category": "技術評価", "name": "Workbench 環境セットアップ", "team": "Joint", "start_date": "2025-10-01", "end_date": "2025-10-31", "deliverable": "環境構築完了", "status": "完了", "is_freeze": False},
    {"wbs": "1.2.1", "phase": "Phase 1", "category": "技術評価", "name": "API 機能調査（Document Translation等）", "team": "Joint", "start_date": "2025-10-01", "end_date": "2025-11-30", "deliverable": "KPMG_Workbench_API一覧", "status": "完了", "is_freeze": False},
    {"wbs": "1.3", "phase": "Phase 1", "category": "課題対応", "name": "手続上の障壁把握", "team": "Joint", "start_date": "2025-10-01", "end_date": "2025-10-31", "deliverable": "課題リスト", "status": "完了", "is_freeze": False},
    {"wbs": "1.3.1", "phase": "Phase 1", "category": "課題対応", "name": "APIアクセス課題解決（12/1完了）", "team": "ALH", "start_date": "2025-11-01", "end_date": "2025-11-30", "deliverable": "解決策ドキュメント", "status": "完了", "is_freeze": False},
    {"wbs": "1.3.2", "phase": "Phase 1", "category": "課題対応", "name": "Global WB Community ローンチ（11/13）", "team": "ALH", "start_date": "2025-11-01", "end_date": "2025-11-30", "deliverable": "Community稼働", "status": "完了", "is_freeze": False},
    {"wbs": "1.4", "phase": "Phase 1", "category": "報告・計画", "name": "Onboarding簡易パワポ作成", "team": "Joint", "start_date": "2025-11-01", "end_date": "2025-11-30", "deliverable": "Onboarding簡易パワポ", "status": "完了", "is_freeze": False},
    {"wbs": "1.4.1", "phase": "Phase 1", "category": "報告・計画", "name": "KC・ALH共同会議", "team": "Joint", "start_date": "2025-11-01", "end_date": "2025-11-30", "deliverable": "N/A", "status": "完了", "is_freeze": False},

    # Phase 2: Onboarding (12-1月) - CSV 22.csv 同期
    # Category: 基盤 (Infrastructure) - 保留
    {"wbs": "2.0.1", "phase": "Phase 2", "category": "基盤", "name": "共同Repository構築", "team": "ALH", "start_date": "2025-12-09", "end_date": "2025-12-10", "deliverable": "GitHub Repo", "status": "未開始", "is_freeze": False},

    # Category: Toolkit (開発環境)
    {"wbs": "2.1", "phase": "Phase 2", "category": "Toolkit", "name": "開発環境セットアップの自動化 (Onboarding Automation)", "team": "Joint", "start_date": "2025-12-09", "end_date": "2025-12-19", "deliverable": "Onboarding Automation Kit", "status": "未開始", "is_freeze": False},

    # Category: Verification (API検証)
    {"wbs": "2.2", "phase": "Phase 2", "category": "Verification", "name": "Workbench API網羅的検証 (API Capability Check)", "team": "KC", "start_date": "2025-12-10", "end_date": "2025-12-25", "deliverable": "API Capability Report", "status": "未開始", "is_freeze": False},

    # Category: Toolkit (開発アセット)
    {"wbs": "2.3", "phase": "Phase 2", "category": "Toolkit", "name": "標準開発アセットの整備 (Dev Starter Kit)", "team": "KC", "start_date": "2025-12-13", "end_date": "2025-12-26", "deliverable": "Dev Starter Kit", "status": "未開始", "is_freeze": False},

    # Category: Platform (ナレッジポータル)
    {"wbs": "2.4", "phase": "Phase 2", "category": "Platform", "name": "セルフサービス型ナレッジポータルの構築", "team": "ALH", "start_date": "2025-12-20", "end_date": "2025-12-26", "deliverable": "Knowledge Portal", "status": "未開始", "is_freeze": False},

    # ★ Freeze Period (年末年始休暇) - 保留
    {"wbs": "-", "phase": "Phase 2", "category": "Freeze", "name": "年末年始休暇 (Freeze Period)", "team": "-", "start_date": "2025-12-28", "end_date": "2026-01-05", "deliverable": "-", "status": "-", "is_freeze": True},

    # Category: Launch (検証・始動)
    {"wbs": "2.5", "phase": "Phase 2", "category": "Launch", "name": "リードタイム検証 & コミュニティ始動", "team": "Joint", "start_date": "2026-01-06", "end_date": "2026-01-31", "deliverable": "1-Week Challenge Report", "status": "未開始", "is_freeze": False},

    # Phase 3: PoC Execution (2-3月) - CSV 22.csv 同期
    # Category: Strategy (計画)
    {"wbs": "3.1", "phase": "Phase 3", "category": "Strategy", "name": "PoCシナリオ選定 & リスク評価 (Planning)", "team": "Joint", "start_date": "2026-02-03", "end_date": "2026-02-14", "deliverable": "PoC Scenario & Risk Assessment", "status": "未開始", "is_freeze": False},

    # Category: Execution (App) - ALH担当
    {"wbs": "3.2", "phase": "Phase 3", "category": "Execution (App)", "name": "PoC 1 (ALH): 多言語ドキュメント処理ツール", "team": "ALH", "start_date": "2026-02-17", "end_date": "2026-03-14", "deliverable": "Document Processing Tool", "status": "未開始", "is_freeze": False},
    {"wbs": "3.3", "phase": "Phase 3", "category": "Execution (App)", "name": "PoC 2 (ALH): 対話型アシスタントUI (Chat UX)", "team": "ALH", "start_date": "2026-02-17", "end_date": "2026-03-14", "deliverable": "Chat Assistant UI", "status": "未開始", "is_freeze": False},

    # Category: Execution (Core) - KC担当
    {"wbs": "3.4", "phase": "Phase 3", "category": "Execution (Core)", "name": "PoC 3 (KC): 社内規程・マニュアル検索 (RAG)", "team": "KC", "start_date": "2026-02-17", "end_date": "2026-03-14", "deliverable": "RAG Search System", "status": "未開始", "is_freeze": False},
    {"wbs": "3.5", "phase": "Phase 3", "category": "Execution (Core)", "name": "PoC 4 (KC): 複雑推論・分析エージェント", "team": "KC", "start_date": "2026-02-17", "end_date": "2026-03-14", "deliverable": "Analysis Agent", "status": "未開始", "is_freeze": False},

    # Category: Decision (成果実演)
    {"wbs": "3.6", "phase": "Phase 3", "category": "Decision", "name": "PoC成果実演・投資判断 (Demo Day)", "team": "Joint", "start_date": "2026-03-17", "end_date": "2026-03-31", "deliverable": "Demo Day & Investment Decision", "status": "未開始", "is_freeze": False},

    # Phase 4: Scaling (4-6月) - CSV 22.csv 同期
    # Category: Standardization (標準化)
    {"wbs": "4.1", "phase": "Phase 4", "category": "Standardization", "name": "開発標準フレームワークの確立", "team": "ALH", "start_date": "2026-04-01", "end_date": "2026-04-18", "deliverable": "KPMG Standard Template", "status": "未開始", "is_freeze": False},

    # Category: Quality (品質保証)
    {"wbs": "4.2", "phase": "Phase 4", "category": "Quality", "name": "品質保証(QA)プロセスの自動化", "team": "KC", "start_date": "2026-04-21", "end_date": "2026-04-30", "deliverable": "QA Automation Process", "status": "未開始", "is_freeze": False},

    # Category: DevOps (デプロイメント)
    {"wbs": "4.3", "phase": "Phase 4", "category": "DevOps", "name": "デプロイメント・パイプラインの構築", "team": "ALH", "start_date": "2026-05-07", "end_date": "2026-05-23", "deliverable": "CI/CD Pipeline", "status": "未開始", "is_freeze": False},
]

# Milestones (including Phase 1) - 新增 Go/No-Go 決定点
MILESTONES = [
    # Phase 1 里程碑
    {"name": "Phase 1 開始", "target_date": "2025-10-01",
     "actual_date": "", "deliverable": "オンボーディング開始", "status": "完了", "type": "milestone"},
    {"name": "Prerequisites認証完了", "target_date": "2025-10-31",
     "actual_date": "", "deliverable": "認証バッジ", "status": "完了", "type": "milestone"},
    {"name": "技術評価レポート完成", "target_date": "2025-11-30",
     "actual_date": "", "deliverable": "評価レポート", "status": "完了", "type": "milestone"},
    # Phase 2 里程碑
    {"name": "Phase 2 開始", "target_date": "2025-12-09",
     "actual_date": "", "deliverable": "チーム拡大", "status": "進行中", "type": "milestone"},
    {"name": "年内素材完備", "target_date": "2025-12-23",
     "actual_date": "", "deliverable": "Cookbook/環境スクリプト", "status": "未開始", "type": "milestone"},
    {"name": "Freeze Period 開始", "target_date": "2025-12-28",
     "actual_date": "", "deliverable": "開発凍結", "status": "未開始", "type": "milestone"},
    {"name": "全員復帰", "target_date": "2026-01-06",
     "actual_date": "", "deliverable": "開発再開", "status": "未開始", "type": "milestone"},
    {"name": "1週間チャレンジ完了", "target_date": "2026-01-14",
     "actual_date": "", "deliverable": "テスト観察ログ", "status": "未開始", "type": "milestone"},
    # ★ Go/No-Go 決定点 (Phase 2)
    {"name": "【Go/No-Go】Phase 2 Review", "target_date": "2026-01-31",
     "actual_date": "", "deliverable": "Phase 2成果レビュー", "status": "未開始", "type": "decision",
     "criteria": "Onboardingサイト稼働, 1週間チャレンジ達成率>80%, Betaテスターフィードバック解決"},
    # Phase 3 里程碑 (PoC Execution)
    {"name": "Phase 3 開始", "target_date": "2026-02-03",
     "actual_date": "", "deliverable": "PoC Execution開始", "status": "未開始", "type": "milestone"},
    {"name": "PoCシナリオ確定", "target_date": "2026-02-14",
     "actual_date": "", "deliverable": "PoC Scenario & Risk Assessment", "status": "未開始", "type": "milestone"},
    {"name": "4 PoC開発完了", "target_date": "2026-03-14",
     "actual_date": "", "deliverable": "4 PoC Prototypes", "status": "未開始", "type": "milestone"},
    # ★ Go/No-Go 決定点 (Phase 3)
    {"name": "【Go/No-Go】Demo Day 投資判断", "target_date": "2026-03-31",
     "actual_date": "", "deliverable": "PoC成果実演・投資判断", "status": "未開始", "type": "decision",
     "criteria": "4 PoC成果比較評価, 製品化優先順位決定, 投資判断完了"},
    # Phase 4 里程碑 (Scaling)
    {"name": "Phase 4 開始", "target_date": "2026-04-01",
     "actual_date": "", "deliverable": "Scaling開始", "status": "未開始", "type": "milestone"},
    {"name": "開発標準フレームワーク完成", "target_date": "2026-04-18",
     "actual_date": "", "deliverable": "KPMG Standard Template", "status": "未開始", "type": "milestone"},
    {"name": "CI/CD Pipeline完成", "target_date": "2026-05-23",
     "actual_date": "", "deliverable": "CI/CD Pipeline", "status": "未開始", "type": "milestone"},
    {"name": "プロジェクト完了", "target_date": "2026-06-30",
     "actual_date": "", "deliverable": "総括レポート", "status": "未開始", "type": "milestone"},
]

# Holiday periods - 新増 Freeze Period
HOLIDAYS = [
    {"team": "KC", "start": "2025-12-24", "end": "2026-01-05", "note": "年末年始休暇"},
    {"team": "ALH", "start": "2025-12-28", "end": "2026-01-05", "note": "年末年始休暇"},
    {"team": "ALL", "start": "2025-12-28", "end": "2026-01-05", "note": "Freeze Period (全員)", "is_freeze": True},
]


def get_next_version(output_dir, base_name):
    """ファイルの次のバージョン番号を取得"""
    pattern = os.path.join(output_dir, f"{base_name}_v*.xlsx")
    existing_files = glob.glob(pattern)

    if not existing_files:
        return 1

    versions = []
    for f in existing_files:
        match = re.search(r'_v(\d+)\.xlsx$', f)
        if match:
            versions.append(int(match.group(1)))

    return max(versions) + 1 if versions else 1


def get_month_year(month):
    """月に基づいて年を返す（10-12月は2025年、1-6月は2026年）"""
    if month >= 10:
        return 2025
    else:
        return 2026


def get_month_last_day(year, month):
    """指定月の最終日を取得"""
    if month == 12:
        return 31
    next_month = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)
    return (next_month - timedelta(days=1)).day


def months_to_dates(months):
    """月配列を開始/終了日付文字列に変換する（後方互換性用）"""
    if not months:
        return None, None

    start_month = min(months)
    end_month = max(months)

    start_year = get_month_year(start_month)
    end_year = get_month_year(end_month)

    start_date = f"{start_year}-{start_month:02d}-01"
    end_day = get_month_last_day(end_year, end_month)
    end_date = f"{end_year}-{end_month:02d}-{end_day:02d}"

    return start_date, end_date


def parse_date(date_str):
    """YYYY-MM-DD形式の日付文字列をdatetimeオブジェクトに変換"""
    if not date_str or date_str == "-":
        return None
    return datetime.strptime(date_str, "%Y-%m-%d")


def date_to_months(start_date_str, end_date_str):
    """開始/終了日付から該当する月のリストを返す（ガントバー描画用）"""
    start = parse_date(start_date_str)
    end = parse_date(end_date_str)
    if not start or not end:
        return []

    months = []
    current = start
    while current <= end:
        month = current.month
        if month not in months:
            months.append(month)
        # 次の月へ
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return months


def month_to_col(month, base_col=11):
    """月をガントチャートの列番号にマッピング
    月順序: 10月, 11月, 12月, 1月, 2月, 3月, 4月, 5月, 6月
    base_col: 最初の月（10月）に対応する列番号
    """
    if month >= 10:
        return base_col + (month - 10)  # 10月=7, 11月=8, 12月=9
    else:
        return base_col + 3 + (month - 1)  # 1月=10, 2月=11, ..., 6月=15


def create_header_style():
    """KPMGブランドのヘッダースタイルを作成"""
    return {
        'font': Font(name='Meiryo UI', bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color=KPMG_BLUE, end_color=KPMG_BLUE, fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_header_style(cell):
    """セルにヘッダースタイルを適用"""
    style = create_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def apply_cell_style(cell, team=None):
    """標準セルスタイルを適用"""
    cell.font = Font(name='Meiryo UI')  # デフォルトフォント
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # チーム列のチームベース色分け
    if team:
        if team == "KC":
            cell.fill = PatternFill(start_color=KC_COLOR, end_color=KC_COLOR, fill_type="solid")
            cell.font = Font(name='Meiryo UI', color="FFFFFF", bold=True)
        elif team == "ALH":
            cell.fill = PatternFill(start_color=ALH_COLOR, end_color=ALH_COLOR, fill_type="solid")
            cell.font = Font(name='Meiryo UI', color="FFFFFF", bold=True)
        elif team == "Joint":
            cell.fill = PatternFill(start_color=JOINT_COLOR, end_color=JOINT_COLOR, fill_type="solid")
            cell.font = Font(name='Meiryo UI', color="FFFFFF", bold=True)


def apply_phase_style(cell, phase):
    """Phase固有のスタイルを適用"""
    color = PHASE_COLORS.get(phase, KPMG_BLUE)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.font = Font(name='Meiryo UI', color="FFFFFF", bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=0)
    cell.border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )


def get_week_start_dates(start_date_str, end_date_str):
    """週の開始日（月曜日）のリストを生成"""
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.strptime(end_date_str, "%Y-%m-%d")

    # 最初の月曜日を探す
    days_until_monday = (7 - start.weekday()) % 7
    if days_until_monday == 0 and start.weekday() != 0:
        days_until_monday = 7
    current = start + timedelta(days=days_until_monday) if start.weekday() != 0 else start

    weeks = []
    while current <= end:
        weeks.append(current)
        current += timedelta(days=7)

    return weeks


def is_in_holiday(date, team):
    """日付がチームの休暇期間内かどうかを確認"""
    for holiday in HOLIDAYS:
        if holiday['team'] == team or team == "Joint":
            h_start = datetime.strptime(holiday['start'], "%Y-%m-%d")
            h_end = datetime.strptime(holiday['end'], "%Y-%m-%d")
            if h_start <= date <= h_end:
                return True
    return False


def create_gantt_sheet(wb):
    """ガントチャートシートを作成（月次ビュー、WBS形式）- 新構造対応"""
    ws = wb.create_sheet("ガントチャート")

    # 月ヘッダー (10月-6月)
    month_labels = ["10月", "11月", "12月", "1月", "2月", "3月", "4月", "5月", "6月"]
    month_values = [10, 11, 12, 1, 2, 3, 4, 5, 6]

    # 新ヘッダー: Phase, WBS, Category, Action, Start, End, Owner, 担当者, Deliverable, Status + 月
    headers = ["Phase", "WBS", "Category", "Action", "Start", "End", "Owner", "担当者", "Deliverable", "Status"]
    base_month_col = len(headers) + 1  # 11列目から月

    # ヘッダーを書き込む
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 月ヘッダーを書き込む
    for i, month_label in enumerate(month_labels):
        col = base_month_col + i
        cell = ws.cell(row=1, column=col, value=month_label)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = 5

    # 列幅を設定（新構造）
    ws.column_dimensions['A'].width = 19  # Phase
    ws.column_dimensions['B'].width = 10  # WBS（三級WBS対応）
    ws.column_dimensions['C'].width = 15  # Category
    ws.column_dimensions['D'].width = 40  # Action
    ws.column_dimensions['E'].width = 13  # Start（新）
    ws.column_dimensions['F'].width = 13  # End（新）
    ws.column_dimensions['G'].width = 10  # Owner
    ws.column_dimensions['H'].width = 10  # 担当者
    ws.column_dimensions['I'].width = 22  # Deliverable
    ws.column_dimensions['J'].width = 10  # Status

    # セル結合のためにPhaseとCategoryでタスクをグループ化
    phase_groups = {}
    category_groups = {}
    for i, task in enumerate(ALL_TASKS):
        phase = task['phase']
        category_key = f"{phase}|{task['category']}"

        if phase not in phase_groups:
            phase_groups[phase] = {'start': i, 'end': i}
        else:
            phase_groups[phase]['end'] = i

        if category_key not in category_groups:
            category_groups[category_key] = {'start': i, 'end': i}
        else:
            category_groups[category_key]['end'] = i

    # タスクを書き込む
    for row, task in enumerate(ALL_TASKS, 2):
        is_freeze = task.get('is_freeze', False)

        # Phase列
        phase_cell = ws.cell(row=row, column=1, value=PHASE_DISPLAY.get(task['phase'], task['phase']))
        apply_phase_style(phase_cell, task['phase'])

        # WBS
        ws.cell(row=row, column=2, value=task['wbs'])
        apply_cell_style(ws.cell(row=row, column=2))

        # Category（旧Sub-Task）
        ws.cell(row=row, column=3, value=task['category'])
        apply_cell_style(ws.cell(row=row, column=3))

        # Action（タスク名）
        ws.cell(row=row, column=4, value=task['name'])
        apply_cell_style(ws.cell(row=row, column=4))

        # Start日付 - 使用 parse_date (line 230) 转换为日期对象
        start_cell = ws.cell(row=row, column=5)
        start_date_obj = parse_date(task.get('start_date', ''))
        start_cell.value = start_date_obj if start_date_obj else task.get('start_date', '')
        start_cell.number_format = 'yyyy/mm/dd'
        apply_cell_style(start_cell)

        # End日付 - 使用 parse_date (line 230) 转换为日期对象
        end_cell = ws.cell(row=row, column=6)
        end_date_obj = parse_date(task.get('end_date', ''))
        end_cell.value = end_date_obj if end_date_obj else task.get('end_date', '')
        end_cell.number_format = 'yyyy/mm/dd'
        apply_cell_style(end_cell)

        # Owner（チーム）
        team_cell = ws.cell(row=row, column=7, value=task['team'])
        apply_cell_style(team_cell, task['team'])

        # 担当者（空欄、下拉選択で選択）
        assignee_cell = ws.cell(row=row, column=8, value="")
        apply_cell_style(assignee_cell)

        # 成果物
        ws.cell(row=row, column=9, value=task['deliverable'])
        apply_cell_style(ws.cell(row=row, column=9))

        # Status列（ガントチャートが主表、下拉選択あり）
        status = task.get('status', '未開始')
        status_cell = ws.cell(row=row, column=10, value=status)
        apply_cell_style(status_cell)
        # ステータスに応じた背景色
        if status in STATUS_COLORS:
            status_cell.fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type="solid")

        # Freeze Period行の特殊スタイル
        if is_freeze:
            for col in range(1, base_month_col + len(month_values)):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=FREEZE_COLOR, end_color=FREEZE_COLOR, fill_type="solid")
                cell.font = Font(name='Meiryo UI', italic=True, color="666666")
            continue  # ガントバー描画をスキップ

        # 月別にガントバーを描画（start_date/end_dateから月を計算）
        task_months = date_to_months(task.get('start_date', ''), task.get('end_date', ''))
        for i, month in enumerate(month_values):
            col = base_month_col + i
            cell = ws.cell(row=row, column=col)

            if month in task_months:
                phase_color = PHASE_COLORS.get(task['phase'], KPMG_BLUE)
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type="solid")

            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

    # Phaseセルを結合
    for phase, indices in phase_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        if start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    # Categoryセルを結合（旧Sub-Task）
    for key, indices in category_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        if start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    # 凡例を追加（Phaseカラー、縦並び）
    legend_row = len(ALL_TASKS) + 4
    ws.cell(row=legend_row, column=1, value="凡例:")
    ws.cell(row=legend_row + 1, column=1, value="Phase 1").fill = PatternFill(start_color=KPMG_DARK_BLUE, end_color=KPMG_DARK_BLUE, fill_type="solid")
    ws.cell(row=legend_row + 1, column=1).font = Font(name='Meiryo UI', color="FFFFFF")
    ws.cell(row=legend_row + 2, column=1, value="Phase 2").fill = PatternFill(start_color=KPMG_BLUE, end_color=KPMG_BLUE, fill_type="solid")
    ws.cell(row=legend_row + 2, column=1).font = Font(name='Meiryo UI', color="FFFFFF")
    ws.cell(row=legend_row + 3, column=1, value="Phase 3").fill = PatternFill(start_color=ALH_COLOR, end_color=ALH_COLOR, fill_type="solid")
    ws.cell(row=legend_row + 3, column=1).font = Font(name='Meiryo UI', color="FFFFFF")
    ws.cell(row=legend_row + 4, column=1, value="Phase 4").fill = PatternFill(start_color=MAGENTA_COLOR, end_color=MAGENTA_COLOR, fill_type="solid")
    ws.cell(row=legend_row + 4, column=1).font = Font(name='Meiryo UI', color="FFFFFF")
    # Freeze Period 凡例
    ws.cell(row=legend_row + 5, column=1, value="Freeze").fill = PatternFill(start_color=FREEZE_COLOR, end_color=FREEZE_COLOR, fill_type="solid")
    ws.cell(row=legend_row + 5, column=1).font = Font(name='Meiryo UI', italic=True, color="666666")

    # 条件付き書式用の行数
    last_row = len(ALL_TASKS) + 1

    # Owner列G（旧E）にデータ検証（下拉選択）を追加
    owner_validation = DataValidation(
        type="list",
        formula1='"KC,ALH,Joint"',  # KC+ALH → Joint に変更
        allow_blank=True
    )
    owner_validation.showDropDown = False  # False = 下拉箭頭を表示
    owner_validation.add(f'G2:G{last_row}')
    ws.add_data_validation(owner_validation)

    # Owner列の条件付き書式（下拉選択後も色が変わるように）
    kc_fill = PatternFill(start_color=KC_COLOR, end_color=KC_COLOR, fill_type="solid")
    alh_fill = PatternFill(start_color=ALH_COLOR, end_color=ALH_COLOR, fill_type="solid")
    joint_fill = PatternFill(start_color=JOINT_COLOR, end_color=JOINT_COLOR, fill_type="solid")
    white_font = Font(name='Meiryo UI', color="FFFFFF", bold=True)

    # Joint を先に判定
    ws.conditional_formatting.add(
        f'G2:G{last_row}',
        FormulaRule(formula=['$G2="Joint"'], fill=joint_fill, font=white_font)
    )
    ws.conditional_formatting.add(
        f'G2:G{last_row}',
        FormulaRule(formula=['$G2="KC"'], fill=kc_fill, font=white_font)
    )
    ws.conditional_formatting.add(
        f'G2:G{last_row}',
        FormulaRule(formula=['$G2="ALH"'], fill=alh_fill, font=white_font)
    )

    # 担当者列H（旧F）にデータ検証（下拉選択）を追加
    assignee_validation = DataValidation(
        type="list",
        formula1='"tsuchiya,tanaka,yang,Ma"',
        allow_blank=True
    )
    assignee_validation.showDropDown = False  # False = 下拉箭頭を表示
    assignee_validation.add(f'H2:H{len(ALL_TASKS) + 1}')
    ws.add_data_validation(assignee_validation)

    # Status列J（旧H）にデータ検証（下拉選択）を追加
    status_validation = DataValidation(
        type="list",
        formula1='"未開始,進行中,完了,ブロック"',
        allow_blank=True
    )
    status_validation.showDropDown = False  # False = 下拉箭頭を表示
    status_validation.add(f'J2:J{last_row}')
    ws.add_data_validation(status_validation)

    # Status列の条件付き書式（下拉選択後も色が変わるように）
    status_gray_fill = PatternFill(start_color=STATUS_COLORS["未開始"], end_color=STATUS_COLORS["未開始"], fill_type="solid")
    status_yellow_fill = PatternFill(start_color=STATUS_COLORS["進行中"], end_color=STATUS_COLORS["進行中"], fill_type="solid")
    status_green_fill = PatternFill(start_color=STATUS_COLORS["完了"], end_color=STATUS_COLORS["完了"], fill_type="solid")
    status_red_fill = PatternFill(start_color=STATUS_COLORS["ブロック"], end_color=STATUS_COLORS["ブロック"], fill_type="solid")

    ws.conditional_formatting.add(
        f'J2:J{last_row}',
        FormulaRule(formula=['$J2="未開始"'], fill=status_gray_fill)
    )
    ws.conditional_formatting.add(
        f'J2:J{last_row}',
        FormulaRule(formula=['$J2="進行中"'], fill=status_yellow_fill)
    )
    ws.conditional_formatting.add(
        f'J2:J{last_row}',
        FormulaRule(formula=['$J2="完了"'], fill=status_green_fill)
    )
    ws.conditional_formatting.add(
        f'J2:J{last_row}',
        FormulaRule(formula=['$J2="ブロック"'], fill=status_red_fill)
    )

    # 条件付き書式：Status="完了"の場合、左側の列をグレーに（Category列Cを除く）
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    # B列(WBS), D-I列(Action, Start, End, Owner, 担当者, Deliverable)
    for col_range in [f'B2:B{last_row}', f'D2:I{last_row}']:
        ws.conditional_formatting.add(
            col_range,
            FormulaRule(
                formula=['$J2="完了"'],
                fill=gray_fill
            )
        )

    # Category列（C列）：該当Category内の全タスクが完了した場合のみグレーに
    for category_key, indices in category_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        # 条件：該当範囲のStatus列（J列）で"完了"の数 = 該当範囲の行数
        formula = f'COUNTIF($J${start_row}:$J${end_row},"完了")=ROWS($J${start_row}:$J${end_row})'
        ws.conditional_formatting.add(
            f'C{start_row}:C{end_row}',
            FormulaRule(formula=[formula], fill=gray_fill)
        )

    # ペインを固定（Status列の後から）
    ws.freeze_panes = 'K2'  # 旧I2→新K2

    return ws


def create_team_sheet(wb, team_name, team_filter):
    """チーム別タスクシートを作成（WBS形式）"""
    ws = wb.create_sheet(f"{team_name}組タスク")

    # WBS形式のヘッダー（新構造：Start/End列追加）
    headers = ["Phase", "WBS", "Category", "Action", "Start", "End", "Months", "Deliverable", "Status"]

    # ヘッダーを書き込む
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 列幅を設定
    col_widths = [18, 8, 18, 42, 12, 12, 15, 22, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # チームでタスクをフィルタ（Freeze Periodを除外、Joint任務は両チームに表示）
    team_tasks = [t for t in ALL_TASKS
                  if (team_filter in t['team'] or t['team'] == 'Joint')
                  and not t.get('is_freeze', False)]

    # Phaseでグループ化（セル結合用）
    phase_groups = {}
    category_groups = {}
    for i, task in enumerate(team_tasks):
        phase = task['phase']
        category_key = f"{phase}|{task['category']}"
        if phase not in phase_groups:
            phase_groups[phase] = {'start': i, 'end': i}
        else:
            phase_groups[phase]['end'] = i
        if category_key not in category_groups:
            category_groups[category_key] = {'start': i, 'end': i}
        else:
            category_groups[category_key]['end'] = i

    # タスクを書き込む
    for row, task in enumerate(team_tasks, 2):
        # Phase列
        phase_cell = ws.cell(row=row, column=1, value=PHASE_DISPLAY.get(task['phase'], task['phase']))
        apply_phase_style(phase_cell, task['phase'])

        # WBS番号
        ws.cell(row=row, column=2, value=task['wbs'])
        apply_cell_style(ws.cell(row=row, column=2))

        # Category（旧Sub-Task）
        ws.cell(row=row, column=3, value=task['category'])
        apply_cell_style(ws.cell(row=row, column=3))

        # Action（タスク名）
        ws.cell(row=row, column=4, value=task['name'])
        apply_cell_style(ws.cell(row=row, column=4))

        # Start日付 - 日期格式 yyyy/mm/dd
        start_cell = ws.cell(row=row, column=5)
        start_date_obj = parse_date(task.get('start_date', ''))
        start_cell.value = start_date_obj if start_date_obj else task.get('start_date', '')
        start_cell.number_format = 'yyyy/mm/dd'
        apply_cell_style(start_cell)

        # End日付 - 日期格式 yyyy/mm/dd
        end_cell = ws.cell(row=row, column=6)
        end_date_obj = parse_date(task.get('end_date', ''))
        end_cell.value = end_date_obj if end_date_obj else task.get('end_date', '')
        end_cell.number_format = 'yyyy/mm/dd'
        apply_cell_style(end_cell)

        # 月（start_date/end_dateから生成）
        months = date_to_months(task.get('start_date'), task.get('end_date'))
        months_str = ", ".join([f"{m}月" for m in months]) if months else ""
        ws.cell(row=row, column=7, value=months_str)
        apply_cell_style(ws.cell(row=row, column=7))

        # 成果物
        ws.cell(row=row, column=8, value=task['deliverable'])
        apply_cell_style(ws.cell(row=row, column=8))

        # ステータス（VLOOKUPでガントチャートから参照）
        total_tasks = len(ALL_TASKS)
        # Status列はガントチャートのJ列（列番号10）なので、B列から9列目
        vlookup_formula = f"=VLOOKUP(B{row},'ガントチャート'!$B$2:$J${total_tasks + 1},9,FALSE)"
        ws.cell(row=row, column=9, value=vlookup_formula)
        apply_cell_style(ws.cell(row=row, column=9))

    # Phaseセルを結合
    for phase, indices in phase_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        if start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    # Categoryセルを結合（旧Sub-Task）
    for _, indices in category_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        if start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    # 条件付き書式：Status="完了"の場合、左側の列をグレーに（Category列Cを除く）
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    last_row = len(team_tasks) + 1
    # B列のみ
    ws.conditional_formatting.add(
        f'B2:B{last_row}',
        FormulaRule(formula=['$I2="完了"'], fill=gray_fill)
    )
    # D-H列（C列Categoryを除外）
    ws.conditional_formatting.add(
        f'D2:H{last_row}',
        FormulaRule(formula=['$I2="完了"'], fill=gray_fill)
    )

    # Category列（C列）：該当Category内の全タスクが完了した場合のみグレーに
    for _, indices in category_groups.items():
        start_row = indices['start'] + 2
        end_row = indices['end'] + 2
        formula = f'COUNTIF($I${start_row}:$I${end_row},"完了")=ROWS($I${start_row}:$I${end_row})'
        ws.conditional_formatting.add(
            f'C{start_row}:C{end_row}',
            FormulaRule(formula=[formula], fill=gray_fill)
        )

    # ペインを固定
    ws.freeze_panes = 'B2'

    return ws


def create_holiday_sheet(wb):
    """休暇カレンダーシートを作成"""
    ws = wb.create_sheet("休暇カレンダー")

    # === 左側：カレンダービュー ===
    title_cell = ws.cell(row=1, column=1, value="カレンダービュー")
    title_cell.font = Font(name='Meiryo UI', bold=True, size=14, color=KPMG_BLUE)

    # === 右側：休暇詳細テーブル（18列目から、3列のみ） ===
    table_start_col = 18  # R列から開始

    # テーブルタイトル
    table_title = ws.cell(row=1, column=table_start_col, value="2024-2025 年末年始休暇カレンダー")
    table_title.font = Font(name='Meiryo UI', bold=True, size=14, color=KPMG_BLUE)
    ws.merge_cells(start_row=1, start_column=table_start_col, end_row=1, end_column=table_start_col + 2)

    # テーブルヘッダー（3列のみ：組織、開始日、終了日）
    headers = ["組織", "開始日", "終了日"]
    for i, header in enumerate(headers):
        cell = ws.cell(row=3, column=table_start_col + i, value=header)
        apply_header_style(cell)

    # テーブルデータ
    for row_idx, holiday in enumerate(HOLIDAYS, 4):
        ws.cell(row=row_idx, column=table_start_col, value=holiday['team'])
        ws.cell(row=row_idx, column=table_start_col + 1, value=holiday['start'])
        ws.cell(row=row_idx, column=table_start_col + 2, value=holiday['end'])

        for i in range(3):
            apply_cell_style(ws.cell(row=row_idx, column=table_start_col + i))

    # テーブル列幅
    ws.column_dimensions[get_column_letter(table_start_col)].width = 8       # 組織
    ws.column_dimensions[get_column_letter(table_start_col + 1)].width = 12  # 開始日
    ws.column_dimensions[get_column_letter(table_start_col + 2)].width = 12  # 終了日

    # === 左側：カレンダー生成 ===

    # 2025年12月と2026年1月のカレンダーを生成
    months = [
        ("2025年12月", 2025, 12),
        ("2026年1月", 2026, 1)
    ]

    start_col = 1
    for month_name, year, month in months:
        ws.cell(row=3, column=start_col, value=month_name).font = Font(name='Meiryo UI', bold=True)

        # 曜日ヘッダー
        week_headers = ["月", "火", "水", "木", "金", "土", "日"]
        for i, header in enumerate(week_headers):
            ws.cell(row=4, column=start_col + i, value=header)

        # 月の最初の日を計算
        first_day = datetime(year, month, 1)
        # 月の日数を取得
        if month == 12:
            days_in_month = 31
        else:
            days_in_month = 31  # 1月

        # カレンダーを埋める
        current_row = 5
        current_col = start_col + first_day.weekday()

        for day in range(1, days_in_month + 1):
            date = datetime(year, month, day)
            cell = ws.cell(row=current_row, column=current_col, value=day)

            # 休暇をチェック
            is_kc_holiday = False
            is_alh_holiday = False

            for holiday in HOLIDAYS:
                h_start = datetime.strptime(holiday['start'], "%Y-%m-%d")
                h_end = datetime.strptime(holiday['end'], "%Y-%m-%d")
                if h_start <= date <= h_end:
                    if holiday['team'] == "KC":
                        is_kc_holiday = True
                    elif holiday['team'] == "ALH":
                        is_alh_holiday = True

            if is_kc_holiday and is_alh_holiday:
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            elif is_kc_holiday:
                cell.fill = PatternFill(start_color=KC_COLOR, end_color=KC_COLOR, fill_type="solid")
                cell.font = Font(name='Meiryo UI', color="FFFFFF")
            elif is_alh_holiday:
                cell.fill = PatternFill(start_color=ALH_COLOR, end_color=ALH_COLOR, fill_type="solid")
                cell.font = Font(name='Meiryo UI', color="FFFFFF")

            current_col += 1
            if current_col > start_col + 6:
                current_col = start_col
                current_row += 1

        start_col += 9  # 次の月へ移動

    # カレンダー列幅を設定
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 5

    # 凡例（縦並び、カレンダー下部に配置）
    legend_row = 12
    ws.cell(row=legend_row, column=1, value="凡例:")

    ws.cell(row=legend_row + 1, column=1, value="KC休暇(Ma)")
    ws.cell(row=legend_row + 1, column=1).fill = PatternFill(start_color=KC_COLOR, end_color=KC_COLOR, fill_type="solid")
    ws.cell(row=legend_row + 1, column=1).font = Font(name='Meiryo UI', color="FFFFFF")

    ws.cell(row=legend_row + 2, column=1, value="ALH休暇")
    ws.cell(row=legend_row + 2, column=1).fill = PatternFill(start_color=ALH_COLOR, end_color=ALH_COLOR, fill_type="solid")
    ws.cell(row=legend_row + 2, column=1).font = Font(name='Meiryo UI', color="FFFFFF")

    ws.cell(row=legend_row + 3, column=1, value="両方休暇")
    ws.cell(row=legend_row + 3, column=1).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # A列を凡例用に広げる
    ws.column_dimensions['A'].width = 12

    return ws


def create_milestone_sheet(wb):
    """マイルストーン追跡シートを作成（Go/No-Go決定点強調付き）"""
    ws = wb.create_sheet("マイルストーン")

    # ヘッダー（新構造：タイプとCriteria列追加）
    headers = ["マイルストーン", "タイプ", "予定日", "実績日", "成果物", "Criteria", "ステータス"]

    # ヘッダーを書き込む
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 列幅を設定
    col_widths = [30, 12, 12, 12, 25, 50, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 決定点スタイル（淡赤色背景）
    decision_fill = PatternFill(start_color=DECISION_COLOR, end_color=DECISION_COLOR, fill_type="solid")
    decision_font = Font(name='Meiryo UI', bold=True, size=10)

    # マイルストーンを書き込む
    for row, milestone in enumerate(MILESTONES, 2):
        is_decision = milestone.get('type') == 'decision'
        type_display = "Go/No-Go" if is_decision else "マイルストーン"

        ws.cell(row=row, column=1, value=milestone['name'])
        ws.cell(row=row, column=2, value=type_display)
        ws.cell(row=row, column=3, value=milestone['target_date'])
        ws.cell(row=row, column=4, value=milestone['actual_date'])
        ws.cell(row=row, column=5, value=milestone['deliverable'])
        ws.cell(row=row, column=6, value=milestone.get('criteria', ''))
        ws.cell(row=row, column=7, value=milestone['status'])

        # 罫線を適用
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            apply_cell_style(cell)

            # Go/No-Go決定点は淡赤色背景＋太字
            if is_decision:
                cell.fill = decision_fill
                cell.font = decision_font

    # ステータスにデータ検証を追加（G列に移動）
    status_validation = DataValidation(
        type="list",
        formula1='"未開始,進行中,完了,遅延"',
        allow_blank=True
    )
    status_validation.add(f'G2:G{len(MILESTONES) + 1}')
    ws.add_data_validation(status_validation)

    # タイプ列にデータ検証を追加
    type_validation = DataValidation(
        type="list",
        formula1='"マイルストーン,Go/No-Go"',
        allow_blank=True
    )
    type_validation.add(f'B2:B{len(MILESTONES) + 1}')
    ws.add_data_validation(type_validation)

    # ペインを固定
    ws.freeze_panes = 'A2'

    return ws


def main():
    """Excelファイルを生成するメイン関数"""
    print("=" * 60)
    print("KPMG Workbench 開発スケジュール Excel 生成ツール v2")
    print("=" * 60)

    # 出力ディレクトリの存在を確認
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "generated_docs", "schedules")
    os.makedirs(output_dir, exist_ok=True)

    # 次のバージョン番号を取得
    base_name = "KPMG_Workbench_Schedule"
    version = get_next_version(output_dir, base_name)

    print(f"\nバージョン番号: v{version}")

    # ワークブックを作成
    wb = Workbook()

    # デフォルトシートを削除
    default_sheet = wb.active
    wb.remove(default_sheet)

    # シートを作成
    print("\n[1/5] ガントチャートを作成中（Phaseセル結合付き）...")
    create_gantt_sheet(wb)

    print("[2/5] KC組タスクビューを作成中...")
    create_team_sheet(wb, "KC", "KC")

    print("[3/5] ALH組タスクビューを作成中...")
    create_team_sheet(wb, "ALH", "ALH")

    print("[4/5] 休暇カレンダーを作成中...")
    create_holiday_sheet(wb)

    print("[5/5] マイルストーン追跡を作成中...")
    create_milestone_sheet(wb)

    # バージョン番号付きでワークブックを保存
    output_filename = f"{base_name}_v{version}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"Excelファイルを生成しました: {output_path}")
    print(f"バージョン: v{version}")
    print("=" * 60)

    # サマリー
    kc_tasks = len([t for t in ALL_TASKS if t['team'] == 'KC'])
    alh_tasks = len([t for t in ALL_TASKS if t['team'] == 'ALH'])
    joint_tasks = len([t for t in ALL_TASKS if t['team'] == 'Joint'])

    print(f"\nタスク統計:")
    print(f"  - KC組タスク: {kc_tasks} 件")
    print(f"  - ALH組タスク: {alh_tasks} 件")
    print(f"  - 連合タスク: {joint_tasks} 件")
    print(f"  - 合計: {len(ALL_TASKS)} 件")
    print(f"  - マイルストーン: {len(MILESTONES)} 件")

    # 自動的にファイルを開く
    import subprocess
    subprocess.run(['open', output_path])

    return output_path


if __name__ == "__main__":
    main()
